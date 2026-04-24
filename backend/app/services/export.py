from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from raijin_shared.models.invoice import Invoice, InvoiceStatus
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

STATUS_LABELS: dict[InvoiceStatus, str] = {
    InvoiceStatus.UPLOADED: "Reçue",
    InvoiceStatus.PROCESSING: "Traitement",
    InvoiceStatus.READY_FOR_REVIEW: "À valider",
    InvoiceStatus.CONFIRMED: "Validée",
    InvoiceStatus.REJECTED: "Rejetée",
    InvoiceStatus.FAILED: "Échec",
}

COLUMNS = [
    ("Date émission", 14),
    ("Date échéance", 14),
    ("Numéro facture", 20),
    ("Fournisseur", 32),
    ("VAT fournisseur", 18),
    ("Total HT", 14),
    ("Total TVA", 14),
    ("Total TTC", 14),
    ("Devise", 8),
    ("Statut", 14),
    ("Confidence OCR", 14),
    ("Fichier source", 36),
    ("ID", 38),
]


def _header_fill() -> PatternFill:
    return PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")


def _safe_decimal(v: Decimal | None) -> float | None:
    if v is None:
        return None
    return float(v)


async def query_invoices_for_export(
    session: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    date_from: date | None = None,
    date_to: date | None = None,
    supplier_id: uuid.UUID | None = None,
    status: InvoiceStatus | None = None,
) -> list[Invoice]:
    stmt = (
        select(Invoice)
        .where(Invoice.tenant_id == tenant_id)
        .options(selectinload(Invoice.supplier))
        .order_by(Invoice.issue_date.desc().nullslast(), Invoice.created_at.desc())
    )
    if date_from is not None:
        stmt = stmt.where(Invoice.issue_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(Invoice.issue_date <= date_to)
    if supplier_id is not None:
        stmt = stmt.where(Invoice.supplier_id == supplier_id)
    if status is not None:
        stmt = stmt.where(Invoice.status == status)

    result = await session.scalars(stmt)
    return list(result.all())


def build_excel(invoices: list[Invoice], *, tenant_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Factures"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")
    for idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = _header_fill()
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Rows
    for row_idx, inv in enumerate(invoices, start=2):
        supplier_name = inv.supplier.name if inv.supplier else ""
        vat = inv.supplier.vat_number if inv.supplier else ""
        ws.cell(row=row_idx, column=1, value=inv.issue_date)
        ws.cell(row=row_idx, column=2, value=inv.due_date)
        ws.cell(row=row_idx, column=3, value=inv.invoice_number or "")
        ws.cell(row=row_idx, column=4, value=supplier_name)
        ws.cell(row=row_idx, column=5, value=vat or "")
        ws.cell(row=row_idx, column=6, value=_safe_decimal(inv.total_ht))
        ws.cell(row=row_idx, column=7, value=_safe_decimal(inv.total_vat))
        ws.cell(row=row_idx, column=8, value=_safe_decimal(inv.total_ttc))
        ws.cell(row=row_idx, column=9, value=inv.currency)
        ws.cell(row=row_idx, column=10, value=STATUS_LABELS.get(inv.status, inv.status.value))
        ws.cell(
            row=row_idx,
            column=11,
            value=_safe_decimal(inv.ocr_confidence),
        )
        ws.cell(row=row_idx, column=12, value=inv.source_file_name)
        ws.cell(row=row_idx, column=13, value=str(inv.id))

        for col in (1, 2):
            ws.cell(row=row_idx, column=col).number_format = "YYYY-MM-DD"
        for col in (6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=11).number_format = "0.00%"

    # Totals row
    total_row = len(invoices) + 3
    if invoices:
        bold = Font(bold=True)
        ws.cell(row=total_row, column=5, value="Totaux").font = bold
        ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row - 2})").font = bold
        ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row - 2})").font = bold
        ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row - 2})").font = bold
        for col in (6, 7, 8):
            ws.cell(row=total_row, column=col).number_format = "#,##0.00"

    # Metadata sheet
    meta = wb.create_sheet("Export")
    meta["A1"] = "Organisation"
    meta["B1"] = tenant_name
    meta["A2"] = "Date d'export"
    meta["B2"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    meta["A3"] = "Factures exportées"
    meta["B3"] = len(invoices)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 40
    for row in (1, 2, 3):
        meta.cell(row=row, column=1).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(tenant_slug: str, *, date_from: date | None, date_to: date | None) -> str:
    today = datetime.utcnow().strftime("%Y%m%d")
    period = ""
    if date_from and date_to:
        period = f"_{date_from.isoformat()}_{date_to.isoformat()}"
    elif date_from:
        period = f"_from_{date_from.isoformat()}"
    elif date_to:
        period = f"_to_{date_to.isoformat()}"
    return f"raijin_{tenant_slug}_{today}{period}.xlsx"
